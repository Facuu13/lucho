import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sqlite3
from datetime import datetime

DB_PATH = "taller.db"
ESTADOS = ["No empezado", "En proceso", "Completado"]

# ------------------ DB ------------------
def init_db():
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    # Tabla base (v1)
    cur.execute("""
    CREATE TABLE IF NOT EXISTS trabajos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        fecha TEXT NOT NULL,
        nombre TEXT NOT NULL,
        patente TEXT NOT NULL,
        telefono TEXT,
        descripcion TEXT
    )
    """)
    conn.commit()
    # Migraciones a v4: agregar columnas si faltan
    _ensure_column(conn, "trabajos", "monto", "REAL")
    _ensure_column(conn, "trabajos", "estado", "TEXT")
    # Setear valores por defecto si hay NULLs
    cur.execute("UPDATE trabajos SET estado = COALESCE(estado, ?) WHERE estado IS NULL", (ESTADOS[0],))
    conn.commit()
    conn.close()

def _ensure_column(conn, table, column, coltype):
    cur = conn.cursor()
    cur.execute(f"PRAGMA table_info({table})")
    cols = [row[1] for row in cur.fetchall()]  # row[1] = name
    if column not in cols:
        cur.execute(f"ALTER TABLE {table} ADD COLUMN {column} {coltype}")
        conn.commit()

def insertar_trabajo(nombre, patente, telefono, descripcion, monto, estado):
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO trabajos (fecha, nombre, patente, telefono, descripcion, monto, estado)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (datetime.now().strftime("%Y-%m-%d %H:%M"), nombre, patente, telefono, descripcion, monto, estado))
    conn.commit()
    conn.close()

def buscar_trabajos(f_patente="", f_telefono="", f_estado="Todos"):
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    query = """SELECT id, fecha, nombre, patente, telefono, descripcion, monto, estado
               FROM trabajos WHERE 1=1"""
    params = []
    if f_patente:
        query += " AND patente LIKE ?"
        params.append(f"%{f_patente.upper()}%")
    if f_telefono:
        query += " AND telefono LIKE ?"
        params.append(f"%{f_telefono}%")
    if f_estado and f_estado != "Todos":
        query += " AND estado = ?"
        params.append(f_estado)
    query += " ORDER BY id DESC"
    cur.execute(query, params)
    rows = cur.fetchall()
    conn.close()
    return rows

def obtener_trabajo_por_id(reg_id: int):
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("""SELECT id, fecha, nombre, patente, telefono, descripcion, monto, estado
                   FROM trabajos WHERE id = ?""", (reg_id,))
    row = cur.fetchone()
    conn.close()
    return row

def actualizar_trabajo(reg_id: int, nombre: str, patente: str, telefono: str, descripcion: str, monto, estado: str):
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("""
        UPDATE trabajos
        SET nombre = ?, patente = ?, telefono = ?, descripcion = ?, monto = ?, estado = ?
        WHERE id = ?
    """, (nombre, patente, telefono, descripcion, monto, estado, reg_id))
    conn.commit()
    conn.close()

def borrar_trabajo(reg_id: int):
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("DELETE FROM trabajos WHERE id = ?", (reg_id,))
    conn.commit()
    conn.close()

# ------------------ UI ------------------
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Registro de Trabajos - Taller (v4)")
        self.geometry("560x520")
        self.resizable(False, False)

        frm = ttk.Frame(self, padding=12)
        frm.pack(fill=tk.BOTH, expand=True)

        ttk.Label(frm, text="Nombre y Apellido *").grid(row=0, column=0, sticky="w", pady=4)
        self.e_nombre = ttk.Entry(frm, width=40)
        self.e_nombre.grid(row=0, column=1, sticky="w")

        ttk.Label(frm, text="Patente *").grid(row=1, column=0, sticky="w", pady=4)
        self.e_patente = ttk.Entry(frm, width=20)
        self.e_patente.grid(row=1, column=1, sticky="w")

        ttk.Label(frm, text="Teléfono").grid(row=2, column=0, sticky="w", pady=4)
        self.e_tel = ttk.Entry(frm, width=20)
        self.e_tel.grid(row=2, column=1, sticky="w")

        ttk.Label(frm, text="Monto ($)").grid(row=3, column=0, sticky="w", pady=4)
        self.e_monto = ttk.Entry(frm, width=20)
        self.e_monto.grid(row=3, column=1, sticky="w")

        ttk.Label(frm, text="Estado").grid(row=4, column=0, sticky="w", pady=4)
        self.cb_estado = ttk.Combobox(frm, values=ESTADOS, state="readonly", width=18)
        self.cb_estado.grid(row=4, column=1, sticky="w")
        self.cb_estado.set(ESTADOS[0])

        ttk.Label(frm, text="Descripción del trabajo").grid(row=5, column=0, sticky="nw", pady=4)
        self.t_desc = tk.Text(frm, width=40, height=8)
        self.t_desc.grid(row=5, column=1, sticky="w")

        btns = ttk.Frame(frm)
        btns.grid(row=6, column=1, sticky="e", pady=12)
        ttk.Button(btns, text="Guardar", command=self.on_guardar).grid(row=0, column=0, padx=4)
        ttk.Button(btns, text="Ver registros", command=self.abrir_lista).grid(row=0, column=1, padx=4)
        ttk.Button(btns, text="Limpiar", command=self.limpiar).grid(row=0, column=2, padx=4)

        ttk.Label(frm, text="(*) Obligatorio", foreground="#666").grid(row=7, column=1, sticky="w")

        self.bind("<Return>", lambda e: self.on_guardar())
        self.e_nombre.focus()

    def on_guardar(self):
        nombre = self.e_nombre.get().strip()
        patente = self.e_patente.get().strip().upper().replace(" ", "")
        telefono = self.e_tel.get().strip().replace(" ", "")
        descripcion = self.t_desc.get("1.0", "end").strip()
        monto_txt = self.e_monto.get().strip().replace(",", ".")
        estado = self.cb_estado.get().strip()

        if not nombre or not patente:
            messagebox.showwarning("Faltan datos", "Nombre y Patente son obligatorios.")
            return

        monto = None
        if monto_txt:
            try:
                monto = float(monto_txt)
            except ValueError:
                messagebox.showwarning("Monto inválido", "Ingresá un monto numérico (ej: 2500 o 1250.50).")
                return

        try:
            insertar_trabajo(nombre, patente, telefono, descripcion, monto, estado)
            messagebox.showinfo("OK", "Trabajo guardado correctamente.")
            self.limpiar()
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo guardar.\n{e}")

    def limpiar(self):
        self.e_nombre.delete(0, "end")
        self.e_patente.delete(0, "end")
        self.e_tel.delete(0, "end")
        self.e_monto.delete(0, "end")
        self.cb_estado.set(ESTADOS[0])
        self.t_desc.delete("1.0", "end")
        self.e_nombre.focus()

    def abrir_lista(self):
        ListaVentana(self)

class ListaVentana(tk.Toplevel):
    COLS = ("id", "fecha", "nombre", "patente", "telefono", "descripcion", "monto", "estado")
    HEADERS = {
        "id": "ID",
        "fecha": "Fecha",
        "nombre": "Nombre y Apellido",
        "patente": "Patente",
        "telefono": "Teléfono",
        "descripcion": "Descripción",
        "monto": "Monto ($)",
        "estado": "Estado"
    }
    WIDTHS = {"id": 50, "fecha": 130, "nombre": 200, "patente": 110, "telefono": 120, "descripcion": 300, "monto": 100, "estado": 120}

    def __init__(self, parent):
        super().__init__(parent)
        self.title("Trabajos guardados")
        self.geometry("1120x560")
        self.resizable(True, True)

        frm_filtros = ttk.Frame(self, padding=(10, 8))
        frm_filtros.pack(fill=tk.X)

        ttk.Label(frm_filtros, text="Patente:").grid(row=0, column=0, padx=4)
        self.f_pat = ttk.Entry(frm_filtros, width=18)
        self.f_pat.grid(row=0, column=1)

        ttk.Label(frm_filtros, text="Teléfono:").grid(row=0, column=2, padx=(12, 4))
        self.f_tel = ttk.Entry(frm_filtros, width=18)
        self.f_tel.grid(row=0, column=3)

        ttk.Label(frm_filtros, text="Estado:").grid(row=0, column=4, padx=(12, 4))
        self.f_estado = ttk.Combobox(frm_filtros, values=["Todos"] + ESTADOS, state="readonly", width=16)
        self.f_estado.grid(row=0, column=5)
        self.f_estado.set("Todos")

        ttk.Button(frm_filtros, text="Buscar", command=self.refrescar).grid(row=0, column=6, padx=8)
        ttk.Button(frm_filtros, text="Limpiar", command=self.limpiar_filtros).grid(row=0, column=7, padx=(0,8))

        export_menu = ttk.Menubutton(frm_filtros, text="Exportar")
        menu = tk.Menu(export_menu, tearoff=0)
        menu.add_command(label="CSV (.csv)", command=self.exportar_csv)
        menu.add_command(label="Excel (.xlsx)", command=self.exportar_excel)
        export_menu["menu"] = menu
        export_menu.grid(row=0, column=8, padx=(12, 0))

        self.tree = ttk.Treeview(self, columns=self.COLS, show="headings", height=18)
        self.tree.pack(fill=tk.BOTH, expand=True, padx=10, pady=8)

        for col in self.COLS:
            self.tree.heading(col, text=self.HEADERS[col])
            anchor = "e" if col == "monto" else "w"
            self.tree.column(col, width=self.WIDTHS[col], anchor=anchor)

        frm_bot = ttk.Frame(self, padding=(10, 6))
        frm_bot.pack(fill=tk.X)
        self.btn_edit = ttk.Button(frm_bot, text="Editar seleccionado", command=self.editar_sel)
        self.btn_del = ttk.Button(frm_bot, text="Borrar seleccionado", command=self.borrar_sel)
        self.btn_edit.pack(side=tk.LEFT, padx=(0, 6))
        self.btn_del.pack(side=tk.LEFT)

        self.refrescar()
        self.tree.bind("<Double-1>", self.on_doble_click)

    # ---------- helpers ----------
    def _rows_actuales(self):
        data = []
        for iid in self.tree.get_children():
            data.append(self.tree.item(iid, "values"))
        return data

    def _default_filename(self, ext):
        fecha = datetime.now().strftime("%Y%m%d_%H%M")
        return f"trabajos_{fecha}.{ext}"

    # ---------- tabla ----------
    def refrescar(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        rows = buscar_trabajos(self.f_pat.get().strip(), self.f_tel.get().strip(), self.f_estado.get())
        for r in rows:
            # Formateo visual del monto para la tabla
            r = list(r)
            monto = r[6]
            r[6] = "" if monto is None else f"{monto:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            self.tree.insert("", "end", values=r)

    def limpiar_filtros(self):
        self.f_pat.delete(0, "end")
        self.f_tel.delete(0, "end")
        self.f_estado.set("Todos")
        self.refrescar()

    def _id_seleccionado(self):
        item = self.tree.focus()
        if not item:
            return None
        vals = self.tree.item(item, "values")
        if not vals:
            return None
        return int(vals[0])

    def on_doble_click(self, _event):
        self.editar_sel()

    def editar_sel(self):
        reg_id = self._id_seleccionado()
        if not reg_id:
            messagebox.showwarning("Seleccioná un registro", "Primero seleccioná una fila.")
            return
        # Ojo: para edición necesitamos los valores reales (sin formato del monto)
        row = obtener_trabajo_por_id(reg_id)
        if not row:
            messagebox.showerror("Error", "No se encontró el registro.")
            return
        EditarVentana(self, reg_id, row, on_saved=self.refrescar)

    def borrar_sel(self):
        reg_id = self._id_seleccionado()
        if not reg_id:
            messagebox.showwarning("Seleccioná un registro", "Primero seleccioná una fila.")
            return
        if messagebox.askyesno("Confirmar borrado", "¿Seguro que querés borrar este registro?"):
            try:
                borrar_trabajo(reg_id)
                self.refrescar()
                messagebox.showinfo("Eliminado", "Registro borrado correctamente.")
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo borrar.\n{e}")

    # ---------- exportar ----------
    def exportar_csv(self):
        rows = self._rows_actuales()
        if not rows:
            messagebox.showinfo("Sin datos", "No hay filas para exportar (probá quitar filtros).")
            return
        filepath = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV", "*.csv")],
            initialfile=self._default_filename("csv"),
            title="Guardar como CSV"
        )
        if not filepath:
            return
        try:
            headers = [self.HEADERS[c] for c in self.COLS]
            with open(filepath, "w", encoding="utf-8", newline="") as f:
                f.write(",".join(headers) + "\n")
                for r in rows:
                    safe = []
                    for cell in r:
                        text = str(cell).replace('"', '""')
                        safe.append(f'"{text}"')
                    f.write(",".join(safe) + "\n")
            messagebox.showinfo("Exportado", f"Exportado correctamente a:\n{filepath}")
        except Exception as e:
            messagebox.showerror("Error al exportar", f"No se pudo exportar CSV.\n{e}")

    def exportar_excel(self):
        rows = self._rows_actuales()
        if not rows:
            messagebox.showinfo("Sin datos", "No hay filas para exportar (probá quitar filtros).")
            return
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
        except ImportError:
            if messagebox.askyesno(
                "openpyxl no encontrado",
                "Para exportar a Excel necesitás el paquete 'openpyxl'.\n"
                "¿Querés exportar a CSV en su lugar?"
            ):
                self.exportar_csv()
            return

        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=self._default_filename("xlsx"),
            title="Guardar como Excel"
        )
        if not filepath:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Trabajos"

            headers = [self.HEADERS[c] for c in self.COLS]
            ws.append(headers)

            for r in rows:
                ws.append(list(r))

            # ancho de columnas aproximado
            for col_idx, header in enumerate(headers, start=1):
                max_len = len(header)
                for row in rows:
                    v = str(row[col_idx - 1]) if row[col_idx - 1] is not None else ""
                    max_len = max(max_len, len(v))
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

            wb.save(filepath)
            messagebox.showinfo("Exportado", f"Exportado correctamente a:\n{filepath}")
        except Exception as e:
            messagebox.showerror("Error al exportar", f"No se pudo exportar Excel.\n{e}")

class EditarVentana(tk.Toplevel):
    def __init__(self, parent, reg_id: int, row_tuple, on_saved=None):
        super().__init__(parent)
        self.title(f"Editar registro #{reg_id}")
        self.geometry("560x520")
        self.resizable(False, False)
        self.reg_id = reg_id
        self.on_saved = on_saved

        # row_tuple = (id, fecha, nombre, patente, telefono, descripcion, monto, estado)
        _, fecha, nombre, patente, telefono, descripcion, monto, estado = row_tuple

        frm = ttk.Frame(self, padding=12)
        frm.pack(fill=tk.BOTH, expand=True)

        ttk.Label(frm, text=f"Fecha (solo lectura): {fecha}", foreground="#555").grid(row=0, column=0, columnspan=2, sticky="w", pady=(0,8))

        ttk.Label(frm, text="Nombre y Apellido *").grid(row=1, column=0, sticky="w", pady=4)
        self.e_nombre = ttk.Entry(frm, width=40); self.e_nombre.grid(row=1, column=1, sticky="w"); self.e_nombre.insert(0, nombre)

        ttk.Label(frm, text="Patente *").grid(row=2, column=0, sticky="w", pady=4)
        self.e_patente = ttk.Entry(frm, width=20); self.e_patente.grid(row=2, column=1, sticky="w"); self.e_patente.insert(0, patente)

        ttk.Label(frm, text="Teléfono").grid(row=3, column=0, sticky="w", pady=4)
        self.e_tel = ttk.Entry(frm, width=20); self.e_tel.grid(row=3, column=1, sticky="w"); self.e_tel.insert(0, telefono if telefono else "")

        ttk.Label(frm, text="Monto ($)").grid(row=4, column=0, sticky="w", pady=4)
        self.e_monto = ttk.Entry(frm, width=20); self.e_monto.grid(row=4, column=1, sticky="w")
        if monto is not None:
            self.e_monto.insert(0, str(monto).replace(".", ","))

        ttk.Label(frm, text="Estado").grid(row=5, column=0, sticky="w", pady=4)
        self.cb_estado = ttk.Combobox(frm, values=ESTADOS, state="readonly", width=18)
        self.cb_estado.grid(row=5, column=1, sticky="w")
        self.cb_estado.set(estado if estado in ESTADOS else ESTADOS[0])

        ttk.Label(frm, text="Descripción del trabajo").grid(row=6, column=0, sticky="nw", pady=4)
        self.t_desc = tk.Text(frm, width=40, height=8); self.t_desc.grid(row=6, column=1, sticky="w")
        if descripcion:
            self.t_desc.insert("1.0", descripcion)

        btns = ttk.Frame(frm); btns.grid(row=7, column=1, sticky="e", pady=12)
        ttk.Button(btns, text="Guardar cambios", command=self.guardar).grid(row=0, column=0, padx=4)
        ttk.Button(btns, text="Cancelar", command=self.destroy).grid(row=0, column=1, padx=4)

        self.bind("<Return>", lambda e: self.guardar())
        self.e_nombre.focus()

    def guardar(self):
        nombre = self.e_nombre.get().strip()
        patente = self.e_patente.get().strip().upper().replace(" ", "")
        telefono = self.e_tel.get().strip().replace(" ", "")
        monto_txt = self.e_monto.get().strip().replace(",", ".")
        estado = self.cb_estado.get().strip()
        descripcion = self.t_desc.get("1.0", "end").strip()

        if not nombre or not patente:
            messagebox.showwarning("Faltan datos", "Nombre y Patente son obligatorios.")
            return

        monto = None
        if monto_txt:
            try:
                monto = float(monto_txt)
            except ValueError:
                messagebox.showwarning("Monto inválido", "Ingresá un monto numérico (ej: 2500 o 1250.50).")
                return

        try:
            actualizar_trabajo(self.reg_id, nombre, patente, telefono, descripcion, monto, estado)
            if self.on_saved:
                self.on_saved()
            messagebox.showinfo("OK", "Cambios guardados correctamente.")
            self.destroy()
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo actualizar.\n{e}")

# ------------------ Main ------------------
if __name__ == "__main__":
    init_db()
    app = App()
    app.mainloop()
